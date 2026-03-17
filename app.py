import streamlit as st
import pandas as pd
import datetime
import json
import os
import streamlit.components.v1 as components  
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

# 💡 [구글 연동 추가] 구글 시트 통신을 위한 라이브러리 추가
import gspread
import google.auth

# 1. 화면 기본 설정
st.set_page_config(page_title="중점추진과제 WBS", layout="wide")

# 💡 상단 여백 최소화 및 [14인치 화면 찌그러짐 방지!] + [대시보드 콤팩트 UI 추가]
st.markdown("""
<style>
    .block-container {
        padding-top: 3.5rem !important; 
        padding-bottom: 1.5rem !important;
        min-width: 1200px !important;  /* 14인치에서도 틀이 무너지지 않게 가로 크기 방어 */
    }
    
    /* 14인치 화면에서 버튼 글씨 두 줄 깨짐 방지 */
    div[data-testid="stButton"] button, 
    div[data-testid="stDownloadButton"] button,
    div[data-testid="stLinkButton"] a {
        white-space: nowrap !important; 
        word-break: keep-all !important; 
        padding-left: 0.3rem !important; 
        padding-right: 0.3rem !important; 
        letter-spacing: -0.3px !important; 
    }
    
    /* 💡 대시보드 카드 콤팩트화 및 상단 버튼들 높이 완벽 통일 */
    div[data-testid="stVerticalBlockBorderWrapper"] {
        padding: 0.8rem !important;
    }
    
    div[data-testid="stButton"] button,
    div[data-testid="stDownloadButton"] button,
    div[data-testid="stLinkButton"] a {
        min-height: 28px !important;
        height: 28px !important;
        padding-top: 2px !important;
        padding-bottom: 2px !important;
        display: inline-flex !important;
        align-items: center !important;
        justify-content: center !important;
    }
    
    div[data-testid="stButton"] button p,
    div[data-testid="stDownloadButton"] button p,
    div[data-testid="stLinkButton"] a p {
        font-size: 13px !important;
        line-height: 1.2 !important;
        margin: 0 !important;
    }
</style>
""", unsafe_allow_html=True)

# 💡 크롬 엉뚱한 자동번역(We->우리는) 완벽 차단용 투명 스크립트!
components.html("""
<script>
    const parentDoc = window.parent.document;
    parentDoc.body.classList.add('notranslate');
    parentDoc.body.setAttribute('translate', 'no');
</script>
""", height=0)

# ==========================================
# 💡 [핵심 엔진] 구글 시트 '멀티 탭(Multi-Tab)' 양방향 연동 엔진!
# ==========================================
SHEET_URL = "https://docs.google.com/spreadsheets/d/1FBvFeUVOy0v2dSJOdZwnrx7aEggBNXkYaAlOGuPx54o/edit"

@st.cache_resource
def get_gspread_client():
    from google.oauth2.service_account import Credentials
    import gspread
    
    # Streamlit Cloud의 보안 저장소(Secrets)에서 구글 키를 불러옵니다.
    creds_dict = dict(st.secrets["gcp_service_account"])
    scopes = [
        "https://www.googleapis.com/auth/cloud-platform",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    return gspread.authorize(credentials)

def load_data():
    try:
        gc = get_gspread_client()
        sh = gc.open_by_url(SHEET_URL)
        all_worksheets = sh.worksheets() 
    except Exception as e:
        st.error(f"⚠️ 구글 시트 연결 실패: {e}")
        return {"projects": {}, "p_members": {}, "categories": {}}

    projects = {}
    p_members = {}
    categories = {}

    for ws in all_worksheets:
        try:
            all_values = ws.get_all_values()
        except:
            continue
            
        if len(all_values) <= 1:
            continue 

        rows = all_values[1:]
        for row in rows:
            row = row + [""] * (12 - len(row))
            p_name, p_mem, cat, sub, man, dep, doc, est_s, est_e, act_s, act_e, memo = row[:12]

            if not p_name: 
                p_name = ws.title 

            if p_name not in projects:
                projects[p_name] = []
                categories[p_name] = []
                p_members[p_name] = p_mem

            if cat and cat not in categories[p_name]:
                categories[p_name].append(cat)

            if sub: 
                projects[p_name].append({
                    "대분류": cat, "소분류": sub, "담당자": man, "부서": dep, "문서": doc, 
                    "예상시작일": est_s, "예상종료일": est_e, 
                    "실제시작일": act_s if act_s else None, 
                    "실제종료일": act_e if act_e else None,
                    "메모": memo 
                })

    return {"projects": projects, "p_members": p_members, "categories": categories}

def save_data(projects, p_members, categories):
    try:
        gc = get_gspread_client()
        sh = gc.open_by_url(SHEET_URL)
        
        existing_ws_titles = [ws.title for ws in sh.worksheets()]
        
        for p_name in projects:
            safe_p_name = p_name[:31] 
            
            export_data = [["프로젝트명", "참여인원", "대분류", "세부업무명", "담당자", "관련부서", "관련문서", "계획시작일", "계획종료일", "실제시작일", "실제종료일", "비고(메모)"]]
            
            p_mem = p_members.get(p_name, "")
            p_cats = categories.get(p_name, [])
            p_tasks = projects.get(p_name, [])
            
            if not p_cats and not p_tasks:
                export_data.append([p_name, p_mem, "", "", "", "", "", "", "", "", "", ""])
            else:
                for cat in p_cats:
                    cat_tasks = [t for t in p_tasks if t['대분류'] == cat]
                    if not cat_tasks:
                        export_data.append([p_name, p_mem, cat, "", "", "", "", "", "", "", "", ""])
                    else:
                        for t in cat_tasks:
                            export_data.append([
                                p_name, p_mem, cat, t.get('소분류',''), t.get('담당자',''), t.get('부서',''), t.get('문서',''),
                                t.get('예상시작일',''), t.get('예상종료일',''),
                                t.get('실제시작일','') or '', t.get('실제종료일','') or '', t.get('메모','')
                            ])
            
            if safe_p_name in existing_ws_titles:
                ws = sh.worksheet(safe_p_name)
            else:
                ws = sh.add_worksheet(title=safe_p_name, rows=max(100, len(export_data)+10), cols=12)
                existing_ws_titles.append(safe_p_name)
                
            ws.clear()
            if export_data:
                ws.update(values=export_data, range_name="A1")
                
    except Exception as e:
        st.error(f"⚠️ 구 시트 저장 실패: {e}")

# --- 앱 실행 시 최초 1회만 구글 시트에서 데이터를 불러옵니다 ---
if 'initialized' not in st.session_state:
    saved_data = load_data()
    st.session_state.projects = saved_data.get("projects", {})
    st.session_state.p_members = saved_data.get("p_members", {})
    st.session_state.categories = saved_data.get("categories", {})
    
    for proj_name, tasks in st.session_state.projects.items():
        if proj_name not in st.session_state.categories:
            cats = list(dict.fromkeys([d['대분류'] for d in tasks if '대분류' in d]))
            st.session_state.categories[proj_name] = cats

    st.session_state.current_page = 'Dashboard' 
    st.session_state.close_dialog = False
    st.session_state.initialized = True

keys_to_init = ['p1_proj', 'p1_mem', 'p1_cat', 'p2_sub', 'p2_man', 'p2_dep', 'p2_doc', 'in_new_cat', 'msg_p1', 'msg_p2']
for k in keys_to_init:
    if k not in st.session_state:
        st.session_state[k] = ""
if 'p2_start' not in st.session_state: st.session_state.p2_start = datetime.date.today()
if 'p2_end' not in st.session_state: st.session_state.p2_end = datetime.date.today()

# ==========================================
# 💡 [핵심 엔진] 3단 달력(타임라인) 및 그리드 생성기
# ==========================================
def get_gantt_assets(min_date, max_date):
    min_date = min_date.replace(day=1)
    if max_date.month == 12:
        max_date = max_date.replace(year=max_date.year+1, month=1, day=1) - datetime.timedelta(days=1)
    else:
        max_date = max_date.replace(month=max_date.month+1, day=1) - datetime.timedelta(days=1)
        
    total_days = max(1, (max_date - min_date).days + 1)
    dates = pd.date_range(min_date, max_date)
    
    years_data = {}
    months_data = []
    
    if len(dates) > 0:
        current_month = dates[0].month
        current_year = dates[0].year
        month_days = 0
        for d in dates:
            years_data[d.year] = years_data.get(d.year, 0) + 1
            if d.month == current_month:
                month_days += 1
            else:
                months_data.append((current_year, current_month, month_days))
                current_month = d.month
                current_year = d.year
                month_days = 1
        months_data.append((current_year, current_month, month_days))
        
    year_html = "<div class='notranslate' style='display:flex; width:100%; text-align:center; font-size:13px; font-weight:bold; color:#2c3e50; border-bottom:1px solid #ccc; background-color:#e8ecef;'>"
    for y, d_count in years_data.items():
        w_pct = (d_count / total_days) * 100
        year_html += f"<div style='width:{w_pct}%; border-right:1px solid #ccc; padding:3px 0;'>{y}년</div>"
    year_html += "</div>"
    
    month_html = "<div class='notranslate' style='display:flex; width:100%; text-align:center; font-size:12px; font-weight:bold; color:#444; border-bottom:1px solid #ccc;'>"
    for y, m, d_count in months_data:
        w_pct = (d_count / total_days) * 100
        month_html += f"<div style='width:{w_pct}%; border-right:1px solid #ccc; padding:3px 0;'>{m}월</div>"
    month_html += "</div>"
    
    week_html = "<div class='notranslate' style='display:flex; width:100%; text-align:center; font-size:10px; color:#666;'>"
    bg_grid_html = "<div style='position:absolute; top:0; left:0; width:100%; height:100%; display:flex; pointer-events:none; z-index:0;'>"
    
    for y, m, d_count in months_data:
        w_pct = (d_count / total_days) * 100
        w_width = 100 / 4
        
        week_html += f"<div style='width:{w_pct}%; display:flex; border-right:1px solid #ccc; box-sizing:border-box;'>"
        for w in range(1, 5):
            br = "border-right:1px dashed #ccc;" if w < 4 else ""
            week_html += f"<div style='width:{w_width}%; {br} padding:2px 0;'>{w}W</div>"
        week_html += "</div>"
        
        bg_grid_html += f"<div style='width:{w_pct}%; display:flex; border-right:1px solid #ccc; box-sizing:border-box;'>"
        for w in range(1, 5):
            br = "border-right:1px dashed #eee;" if w < 4 else ""
            bg_grid_html += f"<div style='width:{w_width}%; {br} height:100%;'></div>"
        bg_grid_html += "</div>"
        
    week_html += "</div>"
    bg_grid_html += "</div>"
    
    timeline_html = f"<div style='width:100%; display:flex; flex-direction:column; background-color:#f4f6f9; border:1px solid #ccc; border-bottom:none; border-radius:4px 4px 0 0;'>{year_html}{month_html}{week_html}</div>"
    
    return timeline_html, bg_grid_html, min_date, max_date, total_days

# ==========================================
# 💡 [팝업] 프로젝트 영구 삭제 (+ 💡 구글 시트 탭 연동 삭제 기능 추가!)
# ==========================================
@st.dialog("프로젝트 영구 삭제")
def delete_popup(p_name):
    st.write(f"**[{p_name}]** 프로젝트를 삭제하시겠습니까?")
    step1 = st.checkbox("네, 삭제를 진행하겠습니다.")
    
    if step1:
        st.error("정말 삭제하시겠습니까? 관련된 모든 세부 업무와 일정이 영구히 삭제됩니다.")
        c1, c2 = st.columns(2)
        if c1.button("예 (영구 삭제)", type="primary", use_container_width=True):
            try:
                gc = get_gspread_client()
                sh = gc.open_by_url(SHEET_URL)
                if len(sh.worksheets()) > 1:
                    ws = sh.worksheet(p_name[:31])
                    sh.del_worksheet(ws)
                else:
                    ws = sh.worksheet(p_name[:31])
                    ws.clear() 
            except:
                pass
                
            if p_name in st.session_state.projects: del st.session_state.projects[p_name]
            if p_name in st.session_state.p_members: del st.session_state.p_members[p_name]
            if p_name in st.session_state.categories: del st.session_state.categories[p_name]
            save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)
            st.success("프로젝트가 삭제되었습니다.")
            st.rerun()
        if c2.button("아니오 (취소)", use_container_width=True):
            st.rerun()

# ==========================================
# 💡 [팝업 1단계] 프로젝트 만들기
# ==========================================
def _save_project_data():
    p_name = st.session_state.p1_proj.strip()
    p_mem = st.session_state.p1_mem.strip()
    cat = st.session_state.p1_cat.strip()
    
    if not p_name:
        st.session_state.msg_p1 = "⚠️ 프로젝트명을 꼭 입력해주세요!"
        return False
        
    if p_name not in st.session_state.projects:
        st.session_state.projects[p_name] = []
        st.session_state.categories[p_name] = []
        
    st.session_state.p_members[p_name] = p_mem
    
    if cat:
        if cat not in st.session_state.categories[p_name]:
            st.session_state.categories[p_name].append(cat)
            st.session_state.msg_p1 = f"✅ 대분류 '{cat}' 추가 완료!"
        else:
            st.session_state.msg_p1 = "⚠️ 이미 존재하는 대분류입니다."
    else:
        st.session_state.msg_p1 = f"✅ 프로젝트 '{p_name}' 생성 완료!"
        
    save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)
    st.session_state.p1_cat = "" 
    return True

def cb_save_p1_continuous():
    _save_project_data()

def cb_save_p1_and_close():
    if _save_project_data():
        st.session_state.msg_p1 = ""
        st.session_state.p1_proj = ""
        st.session_state.p1_mem = ""
        st.session_state.p1_cat = ""
        st.session_state.close_dialog = True

def cb_close_p1():
    st.session_state.msg_p1 = ""
    st.session_state.p1_proj = ""
    st.session_state.p1_mem = ""
    st.session_state.p1_cat = ""
    st.session_state.close_dialog = True

@st.dialog("프로젝트 만들기", width="normal")
def popup_step1():
    if st.session_state.close_dialog:
        st.session_state.close_dialog = False
        st.rerun()
        
    st.write("프로젝트를 생성하고 굵직한 **대분류(뼈대)**를 먼저 만들어 둡니다.")
    
    c1, c2 = st.columns([6, 4])
    p_name = c1.text_input("프로젝트명", key="p1_proj", placeholder="예: 프로젝트 A")
    c2.text_input("참여 인원", key="p1_mem", placeholder="예: 5명")
    
    st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
    st.text_input("추가할 대분류 이름", key="p1_cat", placeholder="예: 1. 기획 / 2. 부품개발 (※ 비워두면 프로젝트만 생성됩니다)")
    
    c_btn1, c_btn2, c_btn3 = st.columns(3)
    c_btn1.button("연속 추가", type="primary", use_container_width=True, on_click=cb_save_p1_continuous)
    c_btn2.button("저장 후 닫기", type="primary", use_container_width=True, on_click=cb_save_p1_and_close)
    c_btn3.button("닫기", use_container_width=True, on_click=cb_close_p1)
        
    if p_name and p_name in st.session_state.categories and st.session_state.categories[p_name]:
        if st.session_state.msg_p1:
            if "⚠️" in st.session_state.msg_p1:
                st.warning(st.session_state.msg_p1)
            else:
                st.success(st.session_state.msg_p1)
                
        st.write(f"**[{p_name}]에 생성된 대분류 목록:**")
        for c in st.session_state.categories[p_name]:
            st.markdown(f"- {c}")
    elif st.session_state.msg_p1:
        if "⚠️" in st.session_state.msg_p1:
            st.warning(st.session_state.msg_p1)
        else:
            st.success(st.session_state.msg_p1)

# ==========================================
# 💡 [새 기능] 프로젝트 기본 정보 변경 엔진 (이름/인원)
# ==========================================
def update_project_info(old_name, new_name, new_mem):
    new_name = new_name.strip()
    new_mem = new_mem.strip()
    
    if not new_name:
        return False, "프로젝트명을 입력해주세요."
        
    # 이름은 그대로, 참여인원만 바뀐 경우
    if old_name == new_name:
        if st.session_state.p_members.get(old_name) != new_mem:
            st.session_state.p_members[old_name] = new_mem
            save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)
        return True, ""
        
    # 이름이 바뀌었는데, 이미 있는 이름인 경우 차단
    if new_name in st.session_state.projects:
        return False, "이미 존재하는 프로젝트명입니다."
        
    # 💡 구글 시트 탭 이름 먼저 변경 시도
    try:
        gc = get_gspread_client()
        sh = gc.open_by_url(SHEET_URL)
        ws = sh.worksheet(old_name[:31])
        ws.update_title(new_name[:31])
    except Exception as e:
        pass # 빈 껍데기 프로젝트라서 아직 탭이 없을 수도 있으므로 패스
        
    # 세션 데이터 이관 작업
    st.session_state.projects[new_name] = st.session_state.projects.pop(old_name)
    st.session_state.categories[new_name] = st.session_state.categories.pop(old_name)
    st.session_state.p_members.pop(old_name, None)
    st.session_state.p_members[new_name] = new_mem
    
    # 만약 현재 상세페이지에서 바꿨다면 현재 페이지 이름도 바꿔줌
    if st.session_state.get('current_page') == old_name:
        st.session_state.current_page = new_name
        
    # 최종 저장 (시트 안의 모든 행에 적힌 프로젝트명이 새 이름으로 덮어써짐)
    save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)
    return True, ""

# ==========================================
# 💡 [팝업 2단계] 세부업무 입력 (프로젝트 수정 기능 추가됨)
# ==========================================
def save_p2(p_sel, cat_sel):
    sub = st.session_state.p2_sub.strip()
    man = st.session_state.p2_man.strip()
    dep = st.session_state.p2_dep.strip()
    doc = st.session_state.p2_doc.strip()
    s_date = st.session_state.p2_start
    e_date = st.session_state.p2_end
        
    if p_sel not in st.session_state.projects: st.session_state.projects[p_sel] = []
    if p_sel not in st.session_state.categories: st.session_state.categories[p_sel] = []
        
    st.session_state.projects[p_sel].append({
        "대분류": cat_sel, "소분류": sub, "담당자": man, "부서": dep, "문서": doc, 
        "예상시작일": str(s_date), "예상종료일": str(e_date), "실제시작일": None, "실제종료일": None,
        "메모": "" 
    })
    
    if cat_sel not in st.session_state.categories[p_sel]:
        st.session_state.categories[p_sel].append(cat_sel)
        
    save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)
    
    if sub: st.session_state.msg_p2 = f"✅ [{cat_sel}] 하위에 '{sub}' 추가 완료!"
    else: st.session_state.msg_p2 = f"✅ [{cat_sel}] 대분류 업무 추가 완료!"
    
    st.session_state.p2_sub = ""
    st.session_state.p2_man = ""
    st.session_state.p2_dep = ""
    st.session_state.p2_doc = ""

def cb_save_continuous(p_sel, cat_sel):
    if cat_sel.strip(): save_p2(p_sel, cat_sel.strip())
    else: st.session_state.msg_p2 = "대분류를 선택하거나 직접 입력해주세요!"

def cb_save_and_close(p_sel, cat_sel):
    if cat_sel.strip(): save_p2(p_sel, cat_sel.strip())
    st.session_state.close_dialog = True 

def cb_just_close():
    st.session_state.msg_p2 = ""
    for k in ['p2_sub', 'p2_man', 'p2_dep', 'p2_doc', 'in_new_cat']:
        if k in st.session_state: st.session_state[k] = ""
    st.session_state.close_dialog = True 

@st.dialog("세부업무 추가 및 프로젝트 수정", width="large")
def popup_step2(p_sel):
    if st.session_state.close_dialog:
        st.session_state.close_dialog = False
        st.rerun()
        
    with st.expander("⚙️ 프로젝트 기본 정보 수정 (이름/참여인원)"):
        st.markdown("<div style='font-size:13px; color:#555; margin-bottom:10px;'>💡 내용 변경 후 아래 버튼을 누르면 구글 시트 탭 이름까지 즉시 변경됩니다.</div>", unsafe_allow_html=True)
        c_e1, c_e2 = st.columns([6, 4])
        new_p_name = c_e1.text_input("프로젝트명 수정", value=p_sel, key=f"edit_pname_{p_sel}")
        new_p_mem = c_e2.text_input("참여인원 수정", value=st.session_state.p_members.get(p_sel, ""), key=f"edit_pmem_{p_sel}")
        
        if st.button("프로젝트 정보 변경 적용", use_container_width=True):
            success, msg = update_project_info(p_sel, new_p_name, new_p_mem)
            if success:
                st.session_state.close_dialog = True
                st.rerun() 
            else:
                st.error(msg)
                
    st.write(f"**[{p_sel}]** 프로젝트에 세부 업무를 추가합니다.")
    
    cat_list = st.session_state.categories.get(p_sel, [])
    options_c = cat_list + ["직접 입력 (새 대분류 추가)"]
    
    c_sel = st.selectbox("대분류 선택", options_c)
    
    if c_sel == "직접 입력 (새 대분류 추가)":
        final_cat = st.text_input("새 대분류 직접 입력", placeholder="여기에 새 대분류를 타이핑하세요", label_visibility="collapsed", key="in_new_cat")
    else:
        final_cat = c_sel
        
    st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
    st.text_input("구분 (세부 업무명)", key="p2_sub", placeholder="예: 도면 설계 (※ 비워두면 대분류 이름만 저장됩니다)")
    
    c3, c4, c5 = st.columns(3)
    c3.text_input("담당자", key="p2_man")
    c4.text_input("관련부서", key="p2_dep")
    c5.text_input("관련문서", key="p2_doc")
    
    c6, c7 = st.columns(2)
    c6.date_input("계획 시작일", key="p2_start")
    c7.date_input("계획 종료일", key="p2_end")
    
    st.write("") 
    c_btn1, c_btn2, c_btn3 = st.columns(3)
    
    c_btn1.button("연속 저장", type="primary", use_container_width=True, on_click=cb_save_continuous, args=(p_sel, final_cat))
    c_btn2.button("저장 후 닫기", type="primary", use_container_width=True, on_click=cb_save_and_close, args=(p_sel, final_cat))
    c_btn3.button("닫기", use_container_width=True, on_click=cb_just_close)
        
    if st.session_state.msg_p2:
        if "⚠️" in st.session_state.msg_p2:
            st.warning(st.session_state.msg_p2)
        else:
            st.success(st.session_state.msg_p2)
            added_tasks = [t['소분류'] for t in st.session_state.projects.get(p_sel, []) if t['대분류'] == final_cat.strip()]
            if added_tasks:
                st.write(f"**[{final_cat.strip()}]에 최근 추가된 업무:**")
                for t in added_tasks[-5:]:
                    disp_t = t if t else "- (대분류만 저장됨)"
                    st.markdown(f"↳ {disp_t}")

# ==========================================
# 💡 [팝업 3] 전체 업무 내용 수정 (+ 꺼짐 방지 & 무한 이동 기능)
# ==========================================
@st.dialog("업무 내용 수정", width="large")
def edit_task_popup(p_name, default_cat=None): 
    raw_data = st.session_state.projects.get(p_name, [])
    cat_list = st.session_state.categories.get(p_name, [])
    
    if not cat_list:
        st.warning("수정할 대분류가 없습니다.")
        return

    st.write("수정할 업무를 선택 후 내용을 덮어쓰거나, **대분류 순서를 위아래로 자유롭게 이동**할 수 있습니다.")
    
    c1, c2 = st.columns(2)
    
    sel_key = f"edit_sel_{p_name}"
    if sel_key not in st.session_state or st.session_state[sel_key] not in cat_list:
        st.session_state[sel_key] = default_cat if default_cat in cat_list else cat_list[0]
        
    current_sel = st.session_state[sel_key]
        
    with c1:
        ui_top = st.container()
        ui_bottom = st.container()
        
        with ui_bottom:
            cc1, cc2 = st.columns(2)
            up_clicked = cc1.button("🔼 위로 이동", use_container_width=True)
            down_clicked = cc2.button("🔽 아래로 이동", use_container_width=True)
            
            if up_clicked:
                idx = st.session_state.categories[p_name].index(current_sel)
                if idx > 0:
                    st.session_state.categories[p_name][idx], st.session_state.categories[p_name][idx-1] = st.session_state.categories[p_name][idx-1], st.session_state.categories[p_name][idx]
                    save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)
            
            if down_clicked:
                idx = st.session_state.categories[p_name].index(current_sel)
                if idx < len(st.session_state.categories[p_name]) - 1:
                    st.session_state.categories[p_name][idx], st.session_state.categories[p_name][idx+1] = st.session_state.categories[p_name][idx+1], st.session_state.categories[p_name][idx]
                    save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)

        with ui_top:
            updated_cats = st.session_state.categories[p_name]
            cat_sel = st.selectbox("1. 대분류 선택 (이동할 대분류)", updated_cats, key=sel_key)
            
            styled_cats = []
            for c in updated_cats:
                if c == cat_sel:
                    styled_cats.append(f"<span style='color:#0056b3; font-weight:bold; border-bottom:2px solid #0056b3;'>{c}</span>")
                else:
                    styled_cats.append(c)
            order_str = " ➔ ".join(styled_cats)
            
            st.markdown(f"""
            <div style='font-size:14px; background:#f4f6f9; padding:10px; border-radius:5px; margin-top:5px; margin-bottom:10px; border:1px solid #ddd;'>
                <b>📌 현재 순서:</b><br>{order_str}
            </div>
            """, unsafe_allow_html=True)

    task_options_with_idx = [(i, d['소분류'] if d['소분류'] else "- (세부업무명 없음)") for i, d in enumerate(raw_data) if d['대분류'] == cat_sel]
    
    with c2:
        if not task_options_with_idx:
            st.info("이 대분류에는 아직 세부 업무가 없습니다. 대분류 이름만 수정할 수 있습니다.")
            with st.form(key=f"edit_empty_cat_{cat_sel}"):
                new_cat_only = st.text_input("빈 대분류 폴더명 수정", value=cat_sel)
                if st.form_submit_button("이름 덮어쓰기", type="primary"):
                    if new_cat_only != cat_sel:
                        if p_name not in st.session_state.categories: st.session_state.categories[p_name] = []
                        idx = st.session_state.categories[p_name].index(cat_sel)
                        st.session_state.categories[p_name][idx] = new_cat_only
                        st.session_state[sel_key] = new_cat_only 
                        save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)
                        st.success("대분류 이름이 변경되었습니다!")
                        st.rerun() 
            return
            
        format_func = lambda x: f"[{x[0]}] {x[1]}"
        selected_tuple = st.selectbox("2. 세부 업무 선택", task_options_with_idx, format_func=format_func)
    
    target_idx = selected_tuple[0]
    target_data = raw_data[target_idx]
    
    st.markdown("<hr style='margin: 10px 0;'>", unsafe_allow_html=True)
    
    with st.form(key=f"edit_form_{target_idx}"):
        f1, f2 = st.columns(2)
        new_cat = f1.text_input("대분류 덮어쓰기 (해당 그룹 일괄 변경)", value=cat_sel)
        new_sub = f2.text_input("세부업무 덮어쓰기", value=target_data['소분류'])
        
        f3, f4, f5 = st.columns(3)
        new_man = f3.text_input("담당자 수정", value=target_data.get('담당자', ''))
        new_dep = f4.text_input("관련부서 수정", value=target_data.get('부서', ''))
        new_doc = f5.text_input("관련문서 수정", value=target_data.get('문서', ''))

        f6, f7 = st.columns(2)
        try:
            est_s = pd.to_datetime(target_data['예상시작일']).date()
            est_e = pd.to_datetime(target_data['예상종료일']).date()
        except:
            est_s = datetime.date.today()
            est_e = datetime.date.today()
            
        new_est_s = f6.date_input("계획 시작일 수정", value=est_s)
        new_est_e = f7.date_input("계획 종료일 수정", value=est_e)
        
        if st.form_submit_button("수정 내용 완벽히 덮어쓰기", type="primary", use_container_width=True):
            if new_cat != cat_sel:
                if p_name not in st.session_state.categories: st.session_state.categories[p_name] = []
                if cat_sel in st.session_state.categories[p_name]:
                    cat_idx = st.session_state.categories[p_name].index(cat_sel)
                    st.session_state.categories[p_name][cat_idx] = new_cat
                for d in st.session_state.projects[p_name]:
                    if d['대분류'] == cat_sel:
                        d['대분류'] = new_cat
                st.session_state[sel_key] = new_cat
            
            target_data['소분류'] = new_sub
            target_data['담당자'] = new_man
            target_data['부서'] = new_dep
            target_data['문서'] = new_doc
            target_data['예상시작일'] = str(new_est_s)
            target_data['예상종료일'] = str(new_est_e)
            
            save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)
            st.success("업무 내용이 깔끔하게 덮어쓰기 되었습니다!")
            st.rerun()

# ==========================================
# 💡 [팝업 4] 현진행일정 업데이트
# ==========================================
def cb_update_progress(p_name, target_idx, sk, ek, memo_val):
    st.session_state.projects[p_name][target_idx]['실제시작일'] = str(st.session_state[sk])
    st.session_state.projects[p_name][target_idx]['실제종료일'] = str(st.session_state[ek])
    st.session_state.projects[p_name][target_idx]['메모'] = memo_val 
    del st.session_state[sk]
    del st.session_state[ek]
    save_data(st.session_state.projects, st.session_state.p_members, st.session_state.categories)
    st.session_state.close_dialog = True 

@st.dialog("진행 현황 입력")
def update_progress_popup(p_name, target_idx):
    if st.session_state.close_dialog:
        st.session_state.close_dialog = False
        st.rerun()
        
    target_data = st.session_state.projects[p_name][target_idx]
    
    disp_sub = target_data['소분류'] if target_data['소분류'] else "-"
    st.write(f"**[{target_data['대분류']}] {disp_sub}**")
    st.info(f"💡 당초 계획 일정: {target_data['예상시작일']} ~ {target_data['예상종료일']}")
    
    memo_val = st.text_input("진행상태 비고/메모", value=target_data.get('메모', ''), placeholder="이슈나 특이사항을 자유롭게 적어주세요")
    st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)

    today = datetime.date.today()
    def parse_date(d_str):
        if d_str: return pd.to_datetime(d_str).date()
        return today

    sk = f"act_s_{target_idx}"
    ek = f"act_e_{target_idx}"
    
    if sk not in st.session_state: st.session_state[sk] = parse_date(target_data.get('실제시작일'))
    if ek not in st.session_state: st.session_state[ek] = parse_date(target_data.get('실제종료일'))

    c6, c7 = st.columns(2)
    c6.date_input("실제 시작일정", key=sk)
    c7.date_input("현재(완료) 일정", key=ek)
    
    st.write("") 
    st.button("저장 및 반영", type="primary", use_container_width=True, on_click=cb_update_progress, args=(p_name, target_idx, sk, ek, memo_val))

# ==========================================
# 💡 [팝업 5] 프로젝트 대분류 진행 요약
# ==========================================
@st.dialog("프로젝트 진행 요약", width="large")
def summary_popup(p_name):
    st.markdown(f"### 📊 [{p_name}] 대분류 일정 요약")
    st.markdown("<div style='font-size:13px; color:#555; margin-bottom:15px;'>※ 막대그래프 (위/회색: 계획 일정 &nbsp;|&nbsp; 아래/파란색: 실적 진행 일정)</div>", unsafe_allow_html=True)
    
    raw_data = st.session_state.projects.get(p_name, [])
    if not raw_data:
        st.info("아직 등록된 업무가 없습니다.")
        return
        
    min_date = pd.to_datetime(min([d['예상시작일'] for d in raw_data]))
    max_date = pd.to_datetime(max([d['예상종료일'] for d in raw_data]))
    for d in raw_data:
        if d.get('실제시작일'): min_date = min(min_date, pd.to_datetime(d['실제시작일']))
        if d.get('실제종료일'): max_date = max(max_date, pd.to_datetime(d['실제종료일']))
        
    timeline_html, bg_grid_html, full_min_date, full_max_date, total_days = get_gantt_assets(min_date, max_date)
    
    st.markdown(f"<div style='width:100%;'>{timeline_html}</div>", unsafe_allow_html=True)
    st.markdown("<hr style='margin: 0; border: none;'>", unsafe_allow_html=True)
    
    categories = st.session_state.categories.get(p_name, [])
    for cat in categories:
        cat_items = [d for d in raw_data if d['대분류'] == cat]
        if not cat_items:
            continue
            
        c_est_starts = [pd.to_datetime(d['예상시작일']) for d in cat_items if d.get('예상시작일')]
        c_est_ends = [pd.to_datetime(d['예상종료일']) for d in cat_items if d.get('예상종료일')]
        c_act_starts = [pd.to_datetime(d['실제시작일']) for d in cat_items if d.get('실제시작일')]
        c_act_ends = [pd.to_datetime(d['실제종료일']) for d in cat_items if d.get('실제종료일')]
        
        c_est_s = min(c_est_starts) if c_est_starts else None
        c_est_e = max(c_est_ends) if c_est_ends else None
        c_act_s = min(c_act_starts) if c_act_starts else None
        c_act_e = max(c_act_ends) if c_act_ends else None
        
        col_cat, col_chart = st.columns([2.5, 7.5], vertical_alignment="center")
        
        with col_cat:
            st.markdown(f"<div class='notranslate' style='padding: 0 10px; background-color: #e9ecef; font-weight: bold; color: #2e5481; font-size: 14px; border-radius: 4px; height: 50px; display: flex; align-items: center; justify-content: center; text-align: center; line-height: 1.2; word-break: keep-all;' title='{cat}'>{cat}</div>", unsafe_allow_html=True)
            
        with col_chart:
            cat_bar_html = f"<div class='notranslate' style='position:relative; width:100%; height:50px; background-color:#fff; border:1px solid #ccc; border-radius:4px; overflow:hidden;'>{bg_grid_html}"
            
            if c_est_s and c_est_e:
                e_l = max(0.0, min(100.0, (c_est_s - full_min_date).days / total_days * 100))
                e_r = max(0.0, min(100.0, ((c_est_e - full_min_date).days + 1) / total_days * 100))
                e_w = max(0.5, e_r - e_l)
                cat_bar_html += f"<div style='position:absolute; top:12px; height:10px; background-color:#8B9BB4; border-radius:4px; left:{e_l}%; width:{e_w}%; z-index:1;' title='계획: {c_est_s.date()} ~ {c_est_e.date()}'></div>"
            
            if c_act_s and c_act_e:
                a_l = max(0.0, min(100.0, (c_act_s - full_min_date).days / total_days * 100))
                a_r = max(0.0, min(100.0, ((c_act_e - full_min_date).days + 1) / total_days * 100))
                a_w = max(0.5, a_r - a_l)
                cat_bar_html += f"<div style='position:absolute; top:28px; height:10px; background-color:#2E5481; border-radius:4px; left:{a_l}%; width:{a_w}%; z-index:2;' title='실적: {c_act_s.date()} ~ {c_act_e.date()}'></div>"
                
            cat_bar_html += "</div>"
            st.markdown(cat_bar_html, unsafe_allow_html=True)
            
        st.markdown("<div style='height: 5px;'></div>", unsafe_allow_html=True)

    st.write("")
    if st.button("닫기", use_container_width=True):
        st.rerun()

# ==========================================
# 💡 [CSS] 무채색 버튼 및 UI 스타일
# ==========================================
st.markdown("""
<style>
    .cat-bar-est { position: absolute; top: 12px; height: 8px; background-color: #A9A9A9; border-radius: 4px; z-index: 1; }
    .cat-bar-act { position: absolute; top: 25px; height: 8px; background-color: #555555; border-radius: 4px; z-index: 2; }
    
    .stButton > button[kind="primary"] {
        background-color: #333333; 
        border-color: #333333;
    }
    .stButton > button[kind="primary"]:hover {
        background-color: #111111; 
        border-color: #111111;
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# 3. 메인 화면 (대시보드 ↔ 상세페이지)
# ==========================================
if st.session_state.current_page == 'Dashboard':
    
    # 💡 [핵심수정] 상단 버튼 너비 비율 조정: 앞의 두 버튼을 1.1로 좁게, 새프로젝트 버튼을 1.8로 넓게 설정
    c_head1, c_head2, c_head3, c_head4 = st.columns([6.0, 1.1, 1.1, 1.8], vertical_alignment="bottom")
    
    with c_head1:
        st.markdown("<h1 class='notranslate' style='margin:0; padding:0; margin-bottom:20px;'>중점추진과제 WBS</h1>", unsafe_allow_html=True)
    
    with c_head2:
        def make_integrated_excel():
            output = BytesIO()
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
                
            if not st.session_state.projects:
                ws = wb.create_sheet(title="프로젝트 없음")
                ws.append(["등록된 프로젝트가 없습니다."])
            else:
                for p_name, project_data in st.session_state.projects.items():
                    safe_p_name = p_name[:31] 
                    ws = wb.create_sheet(title=safe_p_name)
                    headers = ["프로젝트명", "참여인원", "대분류", "세부업무명", "담당자", "관련부서", "관련문서", "계획시작", "계획종료", "실제시작", "실제종료", "비고(메모)"]
                    ws.append(headers)
                    
                    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                    for col in range(1, 13):
                        ws.cell(row=1, column=col).fill = header_fill
                        ws.cell(row=1, column=col).font = Font(bold=True)
                    
                    p_mem = st.session_state.p_members.get(p_name, '')
                    for row in project_data:
                        a_start = str(row['실제시작일']) if row['실제시작일'] else ""
                        a_end = str(row['실제종료일']) if row['실제종료일'] else ""
                        memo = row.get('메모', '')
                        ws.append([p_name, p_mem, row['대분류'], row['소분류'], row['담당자'], row['부서'], row['문서'], str(row['예상시작일']), str(row['예상종료일']), a_start, a_end, memo])
            wb.save(output)
            return output.getvalue()
            
        st.download_button(
            label="📥 엑셀 다운로드",
            data=make_integrated_excel(), 
            file_name=f"WBS_전체일정_{datetime.date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

    with c_head3:
        st.link_button("📊 구글시트 열기", url=SHEET_URL, use_container_width=True)

    with c_head4:
        if st.button("새 프로젝트 만들기", type="primary", use_container_width=True):
            popup_step1()
            
    st.markdown("<hr style='margin: -16px 0 20px 0; border: 1.5px solid #ccc;'>", unsafe_allow_html=True)
    
    st.write("### 프로젝트 요약 대시보드")
    
    project_names = list(st.session_state.projects.keys())
    
    if not project_names:
        st.info("아직 생성된 프로젝트가 없습니다. 위 버튼을 눌러 첫 프로젝트 시작하세요!")
    else:
        cols = st.columns(3) 
        
        for idx, p in enumerate(project_names):
            col_target = cols[idx % 3] 
            
            with col_target:
                with st.container(border=True):
                    raw_data = st.session_state.projects[p]
                    p_mem = st.session_state.p_members.get(p, '')
                    
                    # 💡 UI 콤팩트화: 버튼 이름 축소 및 비율 조정
                    c_title, c_add, c_sum, c_del = st.columns([5.2, 1.6, 1.6, 1.6], vertical_alignment="center")
                    with c_title:
                        if st.button(f"{p}", key=f"go_{p}", use_container_width=True, type="primary"):
                            st.session_state.current_page = p
                            st.rerun()
                    with c_add:
                        if st.button("세부", key=f"add_{p}", use_container_width=True):
                            popup_step2(p) 
                    with c_sum:
                        if st.button("요약", key=f"sum_{p}", use_container_width=True):
                            summary_popup(p)
                    with c_del:
                        if st.button("삭제", key=f"del_{p}", use_container_width=True):
                            delete_popup(p)
                            
                    # 💡 담당자를 버튼 바로 밑으로 바짝 올리고 '업무 00개' 문구 삭제
                    st.markdown(f"<div class='notranslate' style='font-size:13px; color:#444; margin-top:-5px; margin-bottom:5px;'>👤 <b>담당자:</b> {p_mem}</div>", unsafe_allow_html=True)

                    if len(raw_data) > 0:
                        min_date = pd.to_datetime(min([d['예상시작일'] for d in raw_data]))
                        max_date = pd.to_datetime(max([d['예상종료일'] for d in raw_data]))
                        for d in raw_data:
                            if d.get('실제시작일'): min_date = min(min_date, pd.to_datetime(d['실제시작일']))
                            if d.get('실제종료일'): max_date = max(max_date, pd.to_datetime(d['실제종료일']))
                            
                        timeline_html, bg_grid_html, full_min_date, full_max_date, total_days = get_gantt_assets(min_date, max_date)
                        
                        all_est_starts = [pd.to_datetime(d['예상시작일']) for d in raw_data]
                        all_est_ends = [pd.to_datetime(d['예상종료일']) for d in raw_data]
                        p_est_start = min(all_est_starts)
                        p_est_end = max(all_est_ends)
                        
                        e_left = max(0.0, min(100.0, (p_est_start - full_min_date).days / total_days * 100))
                        e_right = max(0.0, min(100.0, ((p_est_end - full_min_date).days + 1) / total_days * 100))
                        e_width = max(0.5, e_right - e_left)
                        
                        # 💡 [최종 핵심 수정] 막대 높이를 10px로 유지하고, top을 '4px'로 조정하여 위/아래 4px씩 완벽한 중앙 정렬 완성!
                        chart_html = f"""
                        <div class="notranslate" style="font-size:12px; color:#777; margin-bottom:4px; line-height:1.2;">📅 {full_min_date.strftime('%y.%m.%d')} ~ {full_max_date.strftime('%y.%m.%d')}</div>
                        <div class="notranslate" style='position:relative; width:100%; height:20px; background-color:#fff; border-radius:4px; border:1px solid #ddd; box-sizing:border-box; overflow:hidden;'>
                            {bg_grid_html}
                            <div style='position:absolute; top:4px; height:10px; background-color:#D3D3D3; border-radius:3px; left:{e_left}%; width:{e_width}%; z-index:1;'></div>
                        """
                        
                        actual_starts = [pd.to_datetime(d['실제시작일']) for d in raw_data if d.get('실제시작일')]
                        actual_ends = [pd.to_datetime(d['실제종료일']) for d in raw_data if d.get('실제종료일')]
                        
                        if actual_starts and actual_ends:
                            p_act_start = min(actual_starts)
                            p_act_end = max(actual_ends)
                            
                            a_left = max(0.0, min(100.0, (p_act_start - full_min_date).days / total_days * 100))
                            a_right = max(0.0, min(100.0, ((p_act_end - full_min_date).days + 1) / total_days * 100))
                            a_width = max(0.5, a_right - a_left)
                            
                            # 실제 진행 막대 역시 top: 4px 로 동일하게 맞춤
                            chart_html += f"<div style='position:absolute; top:4px; height:10px; background-color:#4169E1; opacity:0.85; border-radius:3px; left:{a_left}%; width:{a_width}%; z-index:2;'></div>"
                            
                        # 카드 하단 여백 추가
                        chart_html += "</div><div style='height: 10px;'></div>"
                        
                        st.markdown(chart_html, unsafe_allow_html=True)
                    else:
                        st.markdown("<div style='font-size:12px; color:#999; margin-bottom:10px;'>등록된 일정이 없습니다.</div>", unsafe_allow_html=True)

else:
    # ----------------------------------------
    # 개별 프로젝트 상세 페이지
    # ----------------------------------------
    p_name = st.session_state.current_page
    
    col_t1, col_t2 = st.columns([8.5, 1.5], vertical_alignment="bottom")
    
    with col_t1:
        st.markdown(f"<h2 class='notranslate' style='margin:0; padding:0; font-size: 40px; line-height: 1.2;'>{p_name} WBS</h2>", unsafe_allow_html=True)
        st.markdown(f"<div class='notranslate' style='font-size:15px; color:#555; margin-bottom: 12px;'><b>담당자/참여인원:</b> {st.session_state.p_members.get(p_name, '')}</div>", unsafe_allow_html=True)
        
    with col_t2:
        st.markdown("<div style='padding-top: 5px;'></div>", unsafe_allow_html=True)
        if st.button("대시보드 복귀", use_container_width=True):
            st.session_state.current_page = 'Dashboard'
            st.rerun()
        if st.button("전체 업무 수정", use_container_width=True):
            if p_name in st.session_state.categories and st.session_state.categories[p_name]:
                if f"edit_sel_{p_name}" not in st.session_state:
                    st.session_state[f"edit_sel_{p_name}"] = st.session_state.categories[p_name][0]
            edit_task_popup(p_name)
    
    raw_data = st.session_state.projects.get(p_name, [])
    
    st.markdown("<hr style='margin:-8px 0 10px 0; border: 1.5px solid #ccc;'>", unsafe_allow_html=True)
    
    if len(raw_data) > 0:
        min_date = pd.to_datetime(min([d['예상시작일'] for d in raw_data]))
        max_date = pd.to_datetime(max([d['예상종료일'] for d in raw_data]))
        
        for d in raw_data:
            if d.get('실제시작일'): min_date = min(min_date, pd.to_datetime(d['실제시작일']))
            if d.get('실제종료일'): max_date = max(max_date, pd.to_datetime(d['실제종료일']))
            
        timeline_html, bg_grid_html, full_min_date, full_max_date, total_days = get_gantt_assets(min_date, max_date)

        st.markdown("""
        <style>
            .header-cell { padding: 8px; font-weight: bold; text-align: center; font-size: 14px; background-color: #f0f2f6; border-radius: 4px; }
            .cat-cell { background-color: #e9ecef; padding: 0 15px; font-weight: bold; color: #2e5481; font-size: 15px; border-radius: 4px; height: 45px; display: flex; align-items: center; }
            .item-cell p, .item-cell-left p { margin-bottom: 0px !important; }
            .item-cell { min-height: 45px; display: flex; align-items: center; justify-content: center; font-size: 14px; color: #333; }
            .item-cell-left { min-height: 45px; display: flex; align-items: center; justify-content: flex-start; padding-left: 15px; font-size: 14px; color: #333; }
            .bar-container { position: relative; width: 100%; height: 45px; background-color: #fff; border: 1px solid #ccc; border-top:none; box-sizing: border-box; overflow: visible !important; }
            .bar-est { position: absolute; top: 9px; height: 12px; background-color: #D3D3D3; border-radius: 4px; z-index: 1; }
            .bar-act { position: absolute; top: 24px; height: 12px; background-color: #4169E1; border-radius: 4px; z-index: 2; }
            hr.row-line { margin: 0; border: none; border-bottom: 1px solid #eee; }
        </style>
        """, unsafe_allow_html=True)

        h1, h2, h3, h4, h5, h6 = st.columns([1.5, 1, 1, 1, 1, 5.5], vertical_alignment="bottom")
        
        h1.markdown("<div class='header-cell notranslate my-sticky-header'>구분</div>", unsafe_allow_html=True)
        h2.markdown("<div class='header-cell notranslate'>담당자</div>", unsafe_allow_html=True)
        h3.markdown("<div class='header-cell notranslate'>관련부서</div>", unsafe_allow_html=True)
        h4.markdown("<div class='header-cell notranslate'>관련문서</div>", unsafe_allow_html=True)
        h5.markdown("<div class='header-cell notranslate'>진행 현황</div>", unsafe_allow_html=True)
        
        h6.markdown(f"<div style='width:100%;'>{timeline_html}</div>", unsafe_allow_html=True)

        st.markdown("<div style='height: 5px;'></div>", unsafe_allow_html=True)

        with st.container(height=650):
            categories = st.session_state.categories.get(p_name, [])
            for cat in categories:
                cat_items_with_idx = [(i, d) for i, d in enumerate(raw_data) if d['대분류'] == cat]
                if not cat_items_with_idx:
                    continue 
                    
                st.markdown("<div style='height: 10px;'></div>", unsafe_allow_html=True)
                
                cat_c1, cat_c2, cat_c3, cat_c4, cat_c5, cat_c6 = st.columns([1.5, 1, 1, 1, 1, 5.5], vertical_alignment="center")
                
                with cat_c1:
                    st.markdown(f"""
                    <div style='position: relative; width: 100%; height: 45px;'>
                        <div class='notranslate' style='position: absolute; top: 0; left: 0; width: 285%; padding: 0 15px; background-color: #e9ecef; font-weight: bold; color: #2e5481; font-size: 15px; border-radius: 4px; height: 45px; display: flex; align-items: center; white-space: nowrap; z-index: 10;'>
                            {cat}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)
                
                with cat_c5:
                    st.markdown("<div style='margin-top: 6px;'></div>", unsafe_allow_html=True)
                    if st.button("수정", key=f"edit_cat_{p_name}_{cat}", use_container_width=True):
                        st.session_state[f"edit_sel_{p_name}"] = cat
                        edit_task_popup(p_name)
                        
                with cat_c6:
                    c_est_s, c_est_e = None, None
                    c_act_s, c_act_e = None, None
                    
                    cat_est_starts = [pd.to_datetime(d['예상시작일']) for i, d in cat_items_with_idx if d.get('예상시작일')]
                    cat_est_ends = [pd.to_datetime(d['예상종료일']) for i, d in cat_items_with_idx if d.get('예상종료일')]
                    cat_act_starts = [pd.to_datetime(d['실제시작일']) for i, d in cat_items_with_idx if d.get('실제시작일')]
                    cat_act_ends = [pd.to_datetime(d['실제종료일']) for i, d in cat_items_with_idx if d.get('실제종료일')]
                    
                    if cat_est_starts and cat_est_ends:
                        c_est_s = min(cat_est_starts)
                        c_est_e = max(cat_est_ends)
                    if cat_act_starts and cat_act_ends:
                        c_act_s = min(cat_act_starts)
                        c_act_e = max(cat_act_ends)
                        
                    cat_bar_html = f"<div class='bar-container notranslate' style='background-color: #f8f9fa; overflow: hidden !important;'>{bg_grid_html}"
                    
                    if c_est_s and c_est_e:
                        e_l = max(0.0, min(100.0, (c_est_s - full_min_date).days / total_days * 100))
                        e_r = max(0.0, min(100.0, ((c_est_e - full_min_date).days + 1) / total_days * 100))
                        e_w = max(0.5, e_r - e_l)
                        cat_bar_html += f"<div style='position:absolute; top:12px; height:8px; background-color:#8B9BB4; border-radius:4px; left:{e_l}%; width:{e_w}%; z-index:1;' title='대분류 전체 계획: {c_est_s.date()} ~ {c_est_e.date()}'></div>"
                        
                    if c_act_s and c_act_e:
                        a_l = max(0.0, min(100.0, (c_act_s - full_min_date).days / total_days * 100))
                        a_r = max(0.0, min(100.0, ((c_act_e - full_min_date).days + 1) / total_days * 100))
                        a_w = max(0.5, a_r - a_l)
                        cat_bar_html += f"<div style='position:absolute; top:25px; height:8px; background-color:#2E5481; border-radius:4px; left:{a_l}%; width:{a_w}%; z-index:2;' title='대분류 전체 실적: {c_act_s.date()} ~ {c_act_e.date()}'></div>"
                        
                    cat_bar_html += "</div>"
                    st.markdown(cat_bar_html, unsafe_allow_html=True)
                
                for actual_idx, row in cat_items_with_idx:
                    man = row['담당자'] if row['담당자'] else "-"
                    dep = row['부서'] if row['부서'] else "-"
                    doc = row['문서'] if row['문서'] else "-"
                    
                    disp_sub = row['소분류'] if row['소분류'] else "-"

                    c1, c2, c3, c4, c5, c6 = st.columns([1.5, 1, 1, 1, 1, 5.5], vertical_alignment="center")
                    
                    c1.markdown(f"<div class='item-cell-left notranslate'>&nbsp;&nbsp;&nbsp;&nbsp;↳ {disp_sub}</div>", unsafe_allow_html=True)
                    c2.markdown(f"<div class='item-cell notranslate'>{man}</div>", unsafe_allow_html=True)
                    c3.markdown(f"<div class='item-cell notranslate'>{dep}</div>", unsafe_allow_html=True)
                    c4.markdown(f"<div class='item-cell notranslate'>{doc}</div>", unsafe_allow_html=True)
                    
                    with c5:
                        st.markdown("<div style='margin-top: 6px;'></div>", unsafe_allow_html=True)
                        if st.button("진행입력", key=f"upd_btn_{p_name}_{actual_idx}", use_container_width=True):
                            update_progress_popup(p_name, actual_idx)
                    
                    est_start = pd.to_datetime(row['예상시작일'])
                    est_end = pd.to_datetime(row['예상종료일'])
                    
                    e_left = max(0.0, min(100.0, (est_start - full_min_date).days / total_days * 100))
                    e_right = max(0.0, min(100.0, ((est_end - full_min_date).days + 1) / total_days * 100))
                    e_width = max(0.5, e_right - e_left)
                    
                    bar_html = f"<div class='bar-container notranslate'>{bg_grid_html}<div class='bar-est' style='left: {e_left}%; width: {e_width}%;' title='계획: {est_start.date()} ~ {est_end.date()}'></div>"
                    
                    memo_left = e_left
                    
                    if row.get('실제시작일') and row.get('실제종료일'):
                        act_start = pd.to_datetime(row['실제시작일'])
                        act_end = pd.to_datetime(row['실제종료일'])
                        
                        a_left = max(0.0, min(100.0, (act_start - full_min_date).days / total_days * 100))
                        a_right = max(0.0, min(100.0, ((act_end - full_min_date).days + 1) / total_days * 100))
                        a_width = max(0.5, a_right - a_left)
                        
                        memo_left = a_left 
                        bar_html += f"<div class='bar-act' style='left: {a_left}%; width: {a_width}%;' title='실적: {act_start.date()} ~ {act_end.date()}'></div>"
                    
                    memo_val = row.get('메모', '').strip()
                    if memo_val:
                        bar_html += f"<div style='position:absolute; top:23px; left:calc({memo_left}% + 5px); font-size:12px; font-weight:bold; color:#000; z-index:10; white-space:nowrap; text-shadow: -1px -1px 0 #fff, 1px -1px 0 #fff, -1px 1px 0 #fff, 1px 1px 0 #fff;'>{memo_val}</div>"
                    
                    bar_html += "</div>"
                    c6.markdown(bar_html, unsafe_allow_html=True)
                    
                    st.markdown("<hr class='row-line'>", unsafe_allow_html=True)

        st.caption("※ 막대그래프 (위/회색: 계획 일정 &nbsp;|&nbsp; 아래/파란색: 현재 진행 일정)")
        st.divider()
